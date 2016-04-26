__author__ = 'schien'

import logging
from optparse import make_option
import sys

from django.core.management.base import CommandError
from django.utils.log import getLogger
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from django.db.models.loading import get_model


logger = logging.getLogger(__name__)

version_regexp = r'(\d+(\.\d+)*)'
logger = getLogger('management_commands')


class Command(BaseCommand):
    args = '-f excelfile -m -b -g'
    option_list = BaseCommand.option_list + (
        make_option("-f", "--excelfile",
                    action="store",  # optional because action defaults to "store"
                    dest="excelfile",
                    help="excel file to import from ", ),
        make_option('-m', '--models',
                    action='store',
                    dest='model_names',
                    help='the comma separated list of models you are importing to'),

        make_option('-a', '--app',
                    action='store',
                    dest='app_label',
                    help='the app you are importing to'),

        # make_option('-t', '--type',
        # action='store',
        # dest='template_type',
        # help='the template type (format) your reading from'),

    )

    def handle(self, *args, **options):
        try:

            self.stdout.write("Staring import from excelfile")
            verbose = int(options.get("verbosity", 1)) > 1

            if options['excelfile'] == None:
                raise CommandError('No excelfile specified')

            logger.info("Staring import of models {0} from {1}".format(options['model_names'], options['excelfile']))

            self.import_models(options['excelfile'], models=options['model_names'], verbose=verbose,
                               app_label=options['app_label'])

        except Exception as e:
            logger.error('Admin Command Error: %s', ' '.join(sys.argv), exc_info=sys.exc_info())
            # Raise the exception again so it gets logged in standard error also.
            raise e

    # def update_model_version(self, excelfile, app_label=None):
    # self.stdout.write('updating app in {0}'.format(app_label))
    #
    # app_class = get_model(app_label, 'App')
    # # delete all other version numbers
    # app_class.objects.all().delete()
    #
    #     matches = re.findall(version_regexp, basename(excelfile))
    #     version = matches[0][0]
    #     self.stdout.write('Updating app model version to {0}'.format(version))
    #     app_class.objects.all().delete()
    #     app, _ = app_class.objects.get_or_create(model_version=version)
    #     app.save()

    def set_cell_value(self, house, name, row, col_idx, verbose=False):
        val = row[col_idx].internal_value
        # if val is not None:
            # if dtype:
            #     val = dtype(val)
        if verbose:
            logger.debug('\t{0} = {1}'.format(name, val))
        setattr(house, name, val)


    def get_headers(self, row):
        names = {}
        for i, cell in enumerate(row):
            if cell.internal_value:
                raw_name = cell.internal_value.lower().strip()
                names[raw_name] = i
        return names


    # def replace_names(self, name):
    #     subs = {
    #         'house id': 'id',
    #         'house date': 'age',
    #         'house type': 'type',
    #         'notes': 'comments',
    #         'final_thoughts': 'final_notes'
    #     }
    #     if name in subs:
    #         return subs[name]
    #     return name


    # def get_type(self, key):
    #     if key in {'day', 'time', 'mapping', 'contact', 'adults', 'children', 'bedrooms', 'age', 'type',
    #                'accessibility', 'booking'}:
    #         return int



    def import_models(self, excelFile, models=False, verbose=False, app_label=None):
        wb = load_workbook(excelFile, use_iterators=True)

        models = models.strip().split(',')

        for model in models:

            logger.info('Importing %s' % model)
            model_sheet = wb.get_sheet_by_name(name=model)



            headers = {}

            for row in model_sheet.iter_rows():

                # skip first row
                if row[0].row == 1:
                    headers = self.get_headers(row)
                    if verbose:
                        self.stdout.write(headers)
                    continue

                if not row[0].internal_value:
                    continue

                id = int(row[headers['id']].internal_value)
                logger.debug('Importing %s with id %s ' % (model, id))
                model_class = get_model(app_label, model)

                instance, created = model_class.objects.get_or_create(pk=id)

                if created:
                    instance.pk = id
                else:
                    assert instance.pk == id

                for name, i in headers.items():
                    if name == 'id':
                        continue
                    self.set_cell_value(instance, name, row, i, verbose=verbose)

                instance.save()





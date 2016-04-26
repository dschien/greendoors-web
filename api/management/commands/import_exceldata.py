from os.path import basename
import re
import time

from pygeocoder import Geocoder
import pygeolib


__author__ = 'schien'
import logging
from optparse import make_option
from django.contrib.auth.models import User
from django.core.management.base import BaseCommand, CommandError

from openpyxl import load_workbook

from api.models import InstalledMeasure, HomeOwnerProfile, MeasureCategory, App
from api.models import Measure
from api.models import House

logger = logging.getLogger(__name__)

version_regexp = r'(\d+(\.\d+)*)'


class Command(BaseCommand):
    args = '-f excelfile -m -b -g'
    option_list = BaseCommand.option_list + (
        make_option("-f", "--excelfile",
                    action="store", # optional because action defaults to "store"
                    dest="excelfile",
                    help="excel file to import from ", ),
        make_option('-m', '--measures',
                    action='store_true',
                    dest='measures',
                    default=False,
                    help='Import measures'),
        make_option('-b', '--houses',
                    action='store_true',
                    dest='houses',
                    default=False,
                    help='Import measures'),
        make_option('-e', '--expire-month',
                    action='store_true',
                    dest='expire_month',
                    default=False,
                    help='Expire houses after the event (to be used a month after): change houses marked for CONTACT_MONTH to CONTACT_NONE; also sets open times to CLOSED.'),
        make_option('-c', '--close',
                    action='store_true',
                    dest='close',
                    default=False,
                    help='Set open flag (Sat/Sun) to closed after the event.'),
        make_option('-g', '--geolookup',
                    action='store_true',
                    dest='geolookup',
                    default=False,
                    help='Perform lookup of lat/long pos by address from google maps'),
        make_option('-o', '--override-email',
                    action='store',
                    dest='email',
                    default=False,
                    help='Override email address'),
    )

    excludeList = ['id', 'in progress', 'new since 2012', 'opt in']

    def handle(self, *args, **options):

        self.stdout.write("Staring import from excelfile")
        if options['excelfile'] == None:
            raise CommandError('No excelfile specified')

        if options['measures']:
            self.stdout.write("Staring import of measures from {0}".format(options['excelfile']))
            self.import_measures(options['excelfile'])

        if options['houses']:
            self.stdout.write("Staring import of houses from {0}".format(options['excelfile']))
            verbose = int(options.get("verbosity", 1)) > 1
            self.import_houses(options['excelfile'], geoLookup=options['geolookup'], verbose=verbose,
                               expire_month=options['expire_month'], close=options['close'], email=options['email'])

    def update_model_version(self, excelfile):
        # delete all other version numbers
        App.objects.all().delete()

        matches = re.findall(version_regexp, basename(excelfile))
        version = matches[0][0]
        self.stdout.write('Updating app model version to {0}'.format(version))
        App.objects.all().delete()
        app, _ = App.objects.get_or_create(model_version=version)
        app.save()

    def import_measures(self, excelfile):
        """

        :param excelfile:
        """

        wb = load_workbook(excelfile, use_iterators=True)
        measures_sheet = wb.get_sheet_by_name(name='Measures')

        for row in measures_sheet.iter_rows():
        # skip first row
            if row[0].row == 1:
                continue

            name = row[2].internal_value
            if name is None:
                continue

            category = MeasureCategory.objects.filter(pk=int(row[1].internal_value))[0]
            pk = int(row[0].internal_value)
            measure, created = Measure.objects.get_or_create(pk=pk, name=name, category=category)

            assert measure.pk == pk

            measure.short = row[3].internal_value
            description = row[4].internal_value
            if description:
                measure.description = description
            color = row[5].internal_value
            if color:
                measure.color = color

            template = row[6].internal_value
            if template:
                measure.report_template = template

            measure.save()
        self.update_model_version(excelfile)


    def isNullable(self, name):
        if name in ['comments']:
            return True
        return False

    def set_cell_value(self, house, name, row, col_idx, dtype=type, verbose=False):
        val = row[col_idx].internal_value
        if val is not None:
            if dtype:
                val = dtype(val)
        if val is None and not self.isNullable(name):
            val = 0
        if verbose:
            self.stdout.write('\t{0} = {1}'.format(name, val))
        setattr(house, name, val)


    def get_headers(self, row):
        names = {}
        for i, cell in enumerate(row):
            if cell.internal_value:
                raw_name = cell.internal_value.lower().strip()
                if not raw_name.lower() in self.excludeList:
                    names[i] = self.replace_names(raw_name)
        return names


    def replace_names(self, name):
        subs = {
            'house id': 'id',
            'open times': 'open',
            'house date': 'age',
            'house type': 'type',
            'notes': 'comments'
        }
        if name in subs:
            return subs[name]
        return name


    def get_type(self, key):
        if key in {'open', 'mapping', 'contact', 'adults', 'children', 'bedrooms', 'age', 'type', 'accessibility'}:
            return int


    def geocodeAddress(self, house):
        try:
            coords = Geocoder(client_id='AIzaSyAcpNzt48F8411wYZncqhGU2XTPxpW9X8Y').geocode(house.address).coordinates
            house.latitude = coords[0]
            house.longitude = coords[1]

            return True
        except pygeolib.GeocoderError as e:
            print e
            print 'trying again'
            return False


    def import_houses(self, excelFile, geoLookup=False, verbose=False, expire_month=False, close=False, email=None):
        wb = load_workbook(excelFile, use_iterators=True)
        house_sheet = wb.get_sheet_by_name(name='House')

        houses = []

        headers = {}

        for row in house_sheet.iter_rows():

            # skip first row
            if row[0].row == 1:
                headers = self.get_headers(row)
                if verbose:
                    self.stdout.write(headers)
                continue

            house_id = int(row[0].internal_value)

            # check if opt-ed in
            if row[1].internal_value == 'No':
                continue

            # check if open
            if int(row[5].internal_value) == 0:
                if verbose:
                    self.stdout.write('skipping house with open type: {0}'.format(row[5].internal_value))

            self.stdout.write('importing house with id {0}'.format(house_id))
            house, created = House.objects.get_or_create(pk=house_id)

            if created:
                house.pk = house_id
            else:
                assert house.pk == house_id

            house_sheet_name = 'House ID %04d' % house.pk

            for i, name in headers.items():
                self.set_cell_value(house, name, row, i, self.get_type(name), verbose=verbose)

            if expire_month and house.contact == House.CONTACT_MONTH:
                # if contact is set to month - overwrite to 0  because current clients don't honor this setting
                self.stdout.write(
                    'Applying contact hack - setting contact value from 1 to 0 for house {0}'.format(house.id))
                house.contact = House.CONTACT_NONE

            house.address = "[" + str(house.pk).lstrip("0") + "] " + house.address

            if close or expire_month:
                house.open = House.OPEN_CLOSED

            try:
                if verbose:
                    self.stdout.write('Reading location from template for house {0}'.format(house.id))
                house.latitude = float(row[3].internal_value)
                house.longitude = float(row[4].internal_value)

            except:
                if geoLookup:
                    self.stdout.write('Performing google geolocation lookup for house {0}'.format(house.id))
                    # prevent api
                    while not self.geocodeAddress(house):
                        time.sleep(1)
                        pass

            # create owner and user
            user, _ = User.objects.get_or_create(username='House_{0}_owner'.format(house.pk))

            # if override email is provided, then ignore template settings
            if email is None:
                email = row[17].internal_value
            user.email = email
            user.save()

            owner = None
            owners = HomeOwnerProfile.objects.filter(user_id__exact=user.pk)

            if len(owners) == 0:
                owner = HomeOwnerProfile(user=user)
                owner.save()
            else:
                owner = owners[0]
            house.owner = owner
            house.save()

            if verbose:
                self.stdout.write(house_sheet_name)

            # delete associated measures
            measure_objects_filter = InstalledMeasure.objects.filter(house=house.pk)
            for measure in measure_objects_filter:
                measure.delete()

            self.create_installed_measures(house, wb.get_sheet_by_name(house_sheet_name), verbose=verbose)

        self.update_model_version(excelFile)


    def create_installed_measures(self, house, sheet, verbose=False):
        """

        :param house:
        :param sheet:
        """

        measure_id_col = 0
        disruption_col = 1
        cost_col = 2
        hh_comments_col = 3
        supplier_col = 4
        supplier_urls_col = 5
        product_col = 6
        product_urls_col = 7

        for row in sheet.iter_rows():
        # skip first row
            if row[0].row == 1:
                continue
                # skip empty entries
            if not row[0].internal_value:
                continue

            id = int(row[measure_id_col].internal_value)
            measureType = Measure.objects.get(pk=id)
            measure, _ = InstalledMeasure.objects.get_or_create(house=house, measure=measureType)

            hh_comments = row[hh_comments_col].internal_value
            if hh_comments:
                measure.report_text = hh_comments

            supplier = row[supplier_col].internal_value
            if supplier:
                measure.supplier = supplier

            product = row[product_col].internal_value
            if supplier:
                measure.product = product

            try:
                disruption = row[disruption_col].internal_value
                if disruption:
                    setattr(measure, 'disruption', disruption)
            except:
                if verbose:
                    self.stdout.write('Empty disruption for measure {0} of house {1}'.format(id, house.id))
            try:

                setattr(measure, 'cost', int(row[cost_col].internal_value))
            except:
                if verbose:
                    self.stdout.write('Empty cost for measure {0} of house {1}'.format(id, house.id))

            measure.save()




__author__ = 'schien'

from os.path import basename
import re
import time
import logging
from optparse import make_option
import sys

from pygeocoder import Geocoder
import pygeolib
from django.contrib.auth.models import User
from django.core.management.base import CommandError
from django.utils.log import getLogger
from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from django.db.models.loading import get_model


logger = logging.getLogger(__name__)

version_regexp = r'(\d+(\.\d+)*)'
logger = getLogger('management_commands')


class Command(BaseCommand):
    args = '-f excelfile -m -b -g'
    option_list = BaseCommand.option_list + (
        make_option("-f", "--excelfile",
                    action="store",  # optional because action defaults to "store"
                    dest="excelfile",
                    help="excel file to import from ", ),
        make_option('-m', '--measures',
                    action='store_true',
                    dest='measures',
                    default=False,
                    help='Import measures'),
        make_option('-b', '--houses',
                    action='store_true',
                    dest='houses',
                    default=False,
                    help='Import measures'),
        make_option('-e', '--expire-month',
                    action='store_true',
                    dest='expire_month',
                    default=False,
                    help='Expire houses after the event (to be used a month after): change houses marked for CONTACT_MONTH to CONTACT_NONE; also sets open times to CLOSED.'),
        make_option('-c', '--close',
                    action='store_true',
                    dest='close',
                    default=False,
                    help='Set open flag (Sat/Sun) to closed after the event.'),
        make_option('-g', '--geolookup',
                    action='store_true',
                    dest='geolookup',
                    default=False,
                    help='Perform lookup of lat/long pos by address from google maps'),
        make_option('-o', '--override-email',
                    action='store',
                    dest='email',
                    help='Override email address'),
        make_option('-a', '--app',
                    action='store',
                    dest='app_label',
                    help='the app you are importing to'),
        make_option('-t', '--type',
                    action='store',
                    dest='template_type',
                    help='the template type (format) your reading from'),

    )

    excludeList = ['id', 'in progress', 'new since 2012', 'opt in']

    def handle(self, *args, **options):
        try:

            self.stdout.write("Staring import from excelfile")
            if options['excelfile'] == None:
                raise CommandError('No excelfile specified')

            if options['measures']:
                self.stdout.write("Staring import of measures from {0}".format(options['excelfile']))
                self.import_measures(options['excelfile'], app_label=options['app_label'])

            if options['houses']:
                self.stdout.write("Staring import of houses from {0}".format(options['excelfile']))
                verbose = int(options.get("verbosity", 1)) > 1

                self.import_houses(options['excelfile'], geoLookup=options['geolookup'], verbose=verbose,
                                   expire_month=options['expire_month'], close=options['close'],
                                   overrideemail=options['email'], app_label=options['app_label'])

        except Exception as e:
            logger.error('Admin Command Error: %s', ' '.join(sys.argv), exc_info=sys.exc_info())
            # Raise the exception again so it gets logged in standard error also.
            raise e

    def update_model_version(self, excelfile, app_label=None):
        self.stdout.write('updating app in {0}'.format(app_label))

        app_class = get_model(app_label, 'App')
        # delete all other version numbers
        app_class.objects.all().delete()

        matches = re.findall(version_regexp, basename(excelfile))
        version = matches[0][0]
        self.stdout.write('Updating app model version to {0}'.format(version))
        app_class.objects.all().delete()
        app, _ = app_class.objects.get_or_create(model_version=version)
        app.save()

    def import_measures(self, excelfile, app_label=None):
        """

        :param excelfile:
        """
        self.stdout.write('Importing measure categories for app')
        wb = load_workbook(excelfile, use_iterators=True)
        measures_sheet = wb.get_sheet_by_name(name='Measures')

        for i, row in enumerate(measures_sheet.iter_rows()):
            self.stdout.write('import row %i' % i)
            # skip first row
            if row[0].row == 1:
                continue

            name = row[2].internal_value
            if name is None:
                continue

            measure_category_class = get_model(app_label, 'MeasureCategory')
            category = measure_category_class.objects.filter(pk=int(row[1].internal_value))[0]
            pk = int(row[0].internal_value)

            measure_class = get_model(app_label, 'Measure')
            measure, created = measure_class.objects.get_or_create(pk=pk, name=name, category=category)

            assert measure.pk == pk

            measure.short = row[3].internal_value
            description = row[4].internal_value
            if description:
                measure.description = description
            color = row[5].internal_value
            if color:
                measure.color = color

            template = row[6].internal_value
            if template:
                measure.report_template = template

            measure.save()
        self.update_model_version(excelfile, app_label=app_label)


    def isNullable(self, name):
        if name in ['comments', 'final_thoughts']:
            return True
        return False

    def set_cell_value(self, house, name, row, col_idx, dtype=type, verbose=False):
        val = row[col_idx].internal_value
        if val is not None:
            if dtype:
                val = dtype(val)
        if val is None and not self.isNullable(name):
            val = 0
        if verbose:
            self.stdout.write('\t{0} = {1}'.format(name, val))
        setattr(house, name, val)


    def get_headers(self, row):
        names = {}
        for i, cell in enumerate(row):
            if cell.internal_value:
                raw_name = cell.internal_value.lower().strip()
                if not raw_name.lower() in self.excludeList:
                    names[i] = self.replace_names(raw_name)
        return names


    def replace_names(self, name):
        subs = {
            'house id': 'id',
            'house date': 'age',
            'house type': 'type',
            'notes': 'comments',
            'final_thoughts': 'final_notes'
        }
        if name in subs:
            return subs[name]
        return name


    def get_type(self, key):
        if key in {'day', 'time', 'mapping', 'contact', 'adults', 'children', 'bedrooms', 'age', 'type',
                   'accessibility', 'booking'}:
            return int


    def geocodeAddress(self, house):
        try:
            coords = Geocoder(client_id='AIzaSyAcpNzt48F8411wYZncqhGU2XTPxpW9X8Y').geocode(house.address).coordinates
            house.latitude = coords[0]
            house.longitude = coords[1]

            return True
        except pygeolib.GeocoderError as e:
            print e
            print 'trying again'
            return False


    def import_houses(self, excelFile, geoLookup=False, verbose=False, expire_month=False, close=False,
                      overrideemail=None, app_label=None):
        wb = load_workbook(excelFile, use_iterators=True)
        house_sheet = wb.get_sheet_by_name(name='House')

        houses = []

        headers = {}

        latitude_column = 3
        longitude_column = 4
        open_type_column = 5
        email_column_index = 19

        for row in house_sheet.iter_rows():

            # skip first row
            if row[0].row == 1:
                headers = self.get_headers(row)
                if verbose:
                    self.stdout.write(headers)
                continue

            if not row[0].internal_value:
                continue

            house_id = int(row[0].internal_value)
            logger.info('Handling house %s ' % house_id)
            house_class = get_model(app_label, 'House')


            # check if opt-ed in
            if row[1].internal_value.lower() == 'no':
                if house_class.objects.filter(pk=house_id).exists():
                    house_class.objects.get(pk=house_id).delete()
                continue

            # check if open
            if not row[open_type_column].internal_value:
                logger.error('No open field set for house %s - assume closed' % house_id)
                continue

            if int(row[open_type_column].internal_value) == 0:
                if verbose:
                    self.stdout.write('skipping house with open type: {0}'.format(row[open_type_column].internal_value))

            self.stdout.write('importing house with id {0}'.format(house_id))
            house, created = house_class.objects.get_or_create(pk=house_id)

            if created:
                house.pk = house_id
            else:
                assert house.pk == house_id

            house_sheet_name = 'House ID %04d' % house.pk

            for i, name in headers.items():
                self.set_cell_value(house, name, row, i, self.get_type(name), verbose=verbose)
                if name.lower() == 'email':
                    email_column_index = i

            if expire_month and house.contact == house_class.CONTACT_MONTH:
                # if contact is set to month - overwrite to 0  because current clients don't honor this setting
                self.stdout.write(
                    'Applying contact hack - setting contact value from 1 to 0 for house {0}'.format(house.id))
                house.contact = house_class.CONTACT_NONE

            house.address = "[" + str(house.pk).lstrip("0") + "] " + house.address

            if close or expire_month:
                house.day = house_class.OPEN_CLOSED

            try:
                if verbose:
                    self.stdout.write('Reading location from template for house {0}'.format(house.id))
                house.latitude = float(row[latitude_column].internal_value)
                house.longitude = float(row[longitude_column].internal_value)

            except:
                if geoLookup:
                    self.stdout.write('Performing google geolocation lookup for house {0}'.format(house.id))
                    # prevent api
                    while not self.geocodeAddress(house):
                        time.sleep(1)
                        pass

            # create owner and user
            # data[:75] + '..') if len(data) > 75 else data
            user, _ = User.objects.get_or_create(
                username='{0}_owner_{1}'.format(app_label[0:10] if len(app_label) > 10 else app_label, house.pk))

            # if override email is provided, then ignore template settings
            email = row[email_column_index].internal_value if row[
                                                                  email_column_index].internal_value is not None else 'green-doors+defaultemail@bristol.ac.uk'
            if overrideemail is not None:
                email = overrideemail

            user.email = email
            user.save()

            owner = None

            home_owner_profile_class = get_model(app_label, 'HomeOwnerProfile')
            owners = home_owner_profile_class.objects.filter(user_id__exact=user.pk)

            if len(owners) == 0:
                owner = home_owner_profile_class(user=user)
                owner.save()
            else:
                owner = owners[0]
            house.owner = owner
            house.save()

            if verbose:
                self.stdout.write(house_sheet_name)

            # delete associated measures

            installed_measure_class = get_model(app_label, 'InstalledMeasure')
            measure_objects_filter = installed_measure_class.objects.filter(house=house.pk)
            for measure in measure_objects_filter:
                print "deleting measure %s" % measure.measure.name
                measure.delete()

            self.create_installed_measures(house, wb.get_sheet_by_name(house_sheet_name), installed_measure_class,
                                           app_label, verbose=verbose)

        self.update_model_version(excelFile, app_label=app_label)


    def create_installed_measures(self, house, sheet, installed_measure_class, app_label, verbose=False):
        """

        :param house:
        :param sheet:
        """

        measure_id_col = 0
        disruption_col = 1
        cost_col = 2
        extra_col = 3
        hh_comments_col = 4
        supplier_col = 5
        supplier_urls_col = 6
        product_col = 7
        product_urls_col = 8

        for row in sheet.iter_rows():
            # skip first row
            if row[0].row == 1:
                continue
                # skip empty entries
            if not row[0].internal_value:
                continue

            id = int(row[measure_id_col].internal_value)
            print "creating measure %s" % id
            measure_class = get_model(app_label, 'Measure')
            if not measure_class.objects.filter(pk=id).exists():
                logger.error("Measure with id %s does not exist - ignoring" % id)
                continue

            measureType = measure_class.objects.get(pk=id)
            measure, _ = installed_measure_class.objects.get_or_create(house=house, measure=measureType)

            hh_comments = row[hh_comments_col].internal_value
            if hh_comments:
                measure.report_text = hh_comments

            extras = row[extra_col].internal_value
            if extras:
                measure.extras = extras

            supplier = row[supplier_col].internal_value
            if supplier:
                supplier = supplier.replace(u"\u2013", "-")
                print "supplier %s" % supplier
                measure.supplier = supplier
            supplier_urls = row[supplier_urls_col].internal_value
            if supplier_urls:
                print "supplier urls %s" % supplier_urls
                measure.supplier_urls_text = supplier_urls

            product = row[product_col].internal_value
            if product:
                product = product.replace(u"\u2013", "-")
                print "product %s" % product
                measure.product = product

            product_urls = row[product_urls_col].internal_value
            if product_urls:
                print "product_urls %s" % product_urls
                measure.product_urls_text = product_urls

            try:
                # stars = row[disruption_col].internal_value
                # if not stars is None:
                # stars = len(stars)
                #     setattr(measure, 'disruption', stars)
                disruption = row[disruption_col].internal_value
                if disruption:
                    setattr(measure, 'disruption', disruption)
            except:
                if verbose:
                    self.stdout.write('Empty disruption for measure {0} of house {1}'.format(id, house.id))
            try:

                setattr(measure, 'cost', int(row[cost_col].internal_value))
            except:
                if verbose:
                    self.stdout.write('Empty cost for measure {0} of house {1}'.format(id, house.id))

            measure.save()


__author__ = 'schien'
import logging
from optparse import make_option
from django.core.management.base import BaseCommand, CommandError

from openpyxl import load_workbook

from api.models import InstalledMeasure
from api.models import Measure
from api.models import House

logger = logging.getLogger(__name__)

version_regexp = r'(\d+(\.\d+)*)'


class Command(BaseCommand):
    args = '-f excelfile'
    option_list = BaseCommand.option_list + (
        make_option("-f", "--excelfile",
                    action="store", # optional because action defaults to "store"
                    dest="excelfile",
                    help="excel file to import from ", ),


    )

    excludeList = ['id', 'in progress', 'new since 2012', 'opt in']

    def handle(self, *args, **options):

        self.stdout.write("Staring import from excelfile")
        if options['excelfile'] == None:
            raise CommandError('No excelfile specified')

        verbose = int(options.get("verbosity", 1)) > 1
        self.import_text(options['excelfile'], verbose=verbose)


    def import_text(self, excelFile, geoLookup=False, verbose=False):
        wb = load_workbook(excelFile, use_iterators=True)
        house_sheet = wb.get_sheet_by_name(name='House')

        houses = []

        headers = {}

        for row in house_sheet.iter_rows():

            # skip first row
            if row[0].row == 1:
                headers = self.get_headers(row)
                if verbose:
                    self.stdout.write(headers)
                continue

            # check if opt-ed in
            if row[1].internal_value == 'No':
                continue

            # check if open
            if int(row[5].internal_value) == 0:
                if verbose:
                    self.stdout.write('skipping house with open type: {0}'.format(row[5].internal_value))
                continue

            house_id = int(row[0].internal_value)
            house, created = House.objects.get_or_create(pk=house_id)

            if created:
                house.pk = house_id
            else:
                assert house.pk == house_id

            house_sheet_name = 'House ID %04d' % house.pk

            text = row[18].internal_value
            house.text = text + ": " + house_id

            emails = row[19].internal_value

            house.save()

            if verbose:
                self.stdout.write(house_sheet_name)

            # delete associated measures
            measure_objects_filter = InstalledMeasure.objects.filter(house=house.pk)
            for measure in measure_objects_filter:
                measure.delete()

            self.create_installed_measures(house, wb.get_sheet_by_name(house_sheet_name), verbose=verbose)

        self.update_model_version(excelFile)


    def create_installed_measures(self, house, sheet, verbose=False):
        """

        :param house:
        :param sheet:
        """
        for row in sheet.iter_rows():
        # skip first row
            if row[0].row == 1:
                continue
                # skip empty entries
            if not row[0].internal_value:
                continue

            id = int(row[0].internal_value)
            measureType = Measure.objects.get(pk=id)
            measure, _ = InstalledMeasure.objects.get_or_create(house=house, measure=measureType)

            try:
                stars = row[1].internal_value
                if not stars is None:
                    stars = len(stars)
                    setattr(measure, 'disruption', stars)
            except:
                if verbose:
                    self.stdout.write('Empty disruption for measure {0} of house {1}'.format(id, house.id))
            try:
                setattr(measure, 'cost', int(row[2].internal_value))
            except:
                if verbose:
                    self.stdout.write('Empty cost for measure {0} of house {1}'.format(id, house.id))

            measure.save()

